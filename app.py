import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config("BÁO CÁO KIỂM TOÁN", layout="wide")
st.title("📊 HỆ THỐNG BÁO CÁO KIỂM TOÁN")

# ======================
# RUN INPUT
# ======================
col1, col2 = st.columns(2)
with col1:
    dvkd = st.text_input("MÃ ĐVKD", value="1205")
with col2:
    nguong_tien = st.number_input(
        "NGƯỠNG GIAO DỊCH LỚN (VND)",
        value=500_000_000,
        step=100_000_000
    )

st.markdown("---")

# ======================
# UPLOAD FILE
# ======================
st.subheader("📂 Upload dữ liệu")

file_chuyentien = st.file_uploader(
    "CHUYỂN TIỀN (Napas / Citad / VCB / Swift)",
    type=["xls", "xlsx"]
)

file_huydong = st.file_uploader(
    "HUY ĐỘNG (HDV CKH / KKH)",
    type=["xls", "xlsx"]
)

file_code_sp = st.file_uploader(
    "CODE SẢN PHẨM",
    type=["xls", "xlsx"]
)

file_vang = st.file_uploader(
    "KINH DOANH VÀNG",
    type=["xls", "xlsx"]
)

file_ngoai_te = st.file_uploader(
    "KINH DOANH NGOẠI TỆ",
    type=["xls", "xlsx"]
)

file_ttqt = st.file_uploader(
    "TTQT (TTQT_1 → TTQT_5)",
    type=["xls", "xlsx"]
)

# ======================
# CORE FUNCTIONS
# ======================
def normalize_chuyentien(df):
    df = df.rename(columns={
        "NGAY_GD": "NGAY_GD",
        "SOL_ID": "SOL_ID",
        "LOAI KH CHUYEN": "LOAI_KH",
        "CIF_KH_CHUYEN": "CIF",
        "SO_TIEN_QUY_DOI_VND": "SOTIEN",
        "LOAI_GIAO_DICH": "KENH",
        "NAM GIAO DICH": "NAM"
    })

    if "NAM" not in df or df["NAM"].isna().all():
        df["NAM"] = pd.to_datetime(df["NGAY_GD"], errors="coerce").dt.year

    df["SOTIEN"] = pd.to_numeric(df["SOTIEN"], errors="coerce").fillna(0)
    return df


def bang_2_ct_giao_dich_lon(df, dvkd, nguong):
    df = normalize_chuyentien(df)

    df = df[
        (df["SOL_ID"].astype(str) == str(dvkd)) &
        (df["KENH"].str.upper() != "SWIFT")
    ]

    rows = []
    for nam in sorted(df["NAM"].dropna().unique()):
        df_nam = df[df["NAM"] == nam]
        tong_kh = df_nam["CIF"].nunique()

        gd_kh = df_nam.groupby("CIF", as_index=False)["SOTIEN"].sum()
        gd_lon = gd_kh[gd_kh["SOTIEN"] >= nguong]

        rows.append({
            "NAM": int(nam),
            "TONG_SO_KH": tong_kh,
            "SO_KH_GD_LON": gd_lon["CIF"].nunique(),
            "TONG_TIEN_GD_LON": gd_lon["SOTIEN"].sum() / 1e9,
            "TY_LE": gd_lon["CIF"].nunique() / tong_kh if tong_kh else 0
        })

    return pd.DataFrame(rows)


def bang_1_ct_trong_nuoc(df, dvkd):
    df = normalize_chuyentien(df)
    df = df[df["SOL_ID"].astype(str) == str(dvkd)]
    df = df[df["KENH"].str.upper() != "SWIFT"]

    kq = (
        df.groupby(["KENH", "NAM"])
        .agg(
            TONG_TIEN=("SOTIEN", "sum"),
            SO_GD=("SOTIEN", "count")
        )
        .reset_index()
    )

    kq["TONG_TIEN"] = kq["TONG_TIEN"] / 1e9
    return kq


def bang_3_ct_nuoc_ngoai(df, dvkd):
    df = normalize_chuyentien(df)
    df = df[
        (df["SOL_ID"].astype(str) == str(dvkd)) &
        (df["KENH"].str.upper() == "SWIFT")
    ]

    return (
        df.groupby(["NAM", "LOAI_GIAO_DICH"])
        .agg(
            TONG_TIEN=("SOTIEN", "sum"),
            SO_GD=("SOTIEN", "count")
        )
        .reset_index()
    )


# ======================
# PROCESS
# ======================
if st.button("⚙️ XỬ LÝ TOÀN BỘ BÁO CÁO"):
    if not file_chuyentien:
        st.error("❌ Chưa upload file CHUYỂN TIỀN")
        st.stop()

    with st.spinner("Đang xử lý dữ liệu..."):
        df_ct = pd.read_excel(file_chuyentien)

        bang1_ct = bang_1_ct_trong_nuoc(df_ct, dvkd)
        bang2_ct = bang_2_ct_giao_dich_lon(df_ct, dvkd, nguong_tien)
        bang3_ct = bang_3_ct_nuoc_ngoai(df_ct, dvkd)

    st.success("✅ Xử lý xong")

    st.subheader("BẢNG 1_CT – TRONG NƯỚC")
    st.dataframe(bang1_ct)

    st.subheader("BẢNG 2_CT – GIAO DỊCH LỚN")
    st.dataframe(bang2_ct)

    st.subheader("BẢNG 3_CT – NƯỚC NGOÀI")
    st.dataframe(bang3_ct)

    # ======================
    # EXPORT EXCEL
    # ======================
    wb = load_workbook("templates/BAO_CAO_TEMPLATE.xlsx")

    wb["BANG 1_CT_TRONG NUOC"].cell(5, 2).value = "AUTO"
    wb["BANG 2_CT_GIAO DICH LON"].cell(5, 2).value = "AUTO"
    wb["BANG 3_CT_NUOC NGOAI"].cell(5, 2).value = "AUTO"

    output = BytesIO()
    wb.save(output)

    st.download_button(
        "⬇️ TẢI FILE BÁO CÁO",
        data=output.getvalue(),
        file_name="BAO_CAO_KIEM_TOAN.xlsx"
    )
