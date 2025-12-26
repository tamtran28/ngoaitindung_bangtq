from openpyxl import load_workbook
from io import BytesIO


def export_excel(result_dict, template_path):
    wb = load_workbook(template_path)

    # ví dụ
    ws = wb["BANG 2_CT_GIAO DICH LON"]
    df = result_dict["BANG_2_CT"]

    start_row = 5
    for _, r in df.iterrows():
        ws.cell(start_row, 3).value = r["TONG_SO_KH"]
        ws.cell(start_row, 4).value = r["TONG_TIEN_GD_LON"]
        ws.cell(start_row, 5).value = r["SO_KH_GD_LON"]
        ws.cell(start_row, 6).value = r["TY_LE"]
        start_row += 1

    output = BytesIO()
    wb.save(output)
    return output
